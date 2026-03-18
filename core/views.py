from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Avg, Q, Sum, Max, Min, Case, When, IntegerField, F
from django.db.models.functions import Coalesce
from django.db.models import Value
from django.core.paginator import Paginator
from django.utils import timezone
from django.urls import reverse
from django.db import transaction
from django.views.decorators.http import require_POST
from datetime import timedelta, datetime
from calendar import monthrange
import json
import re
from .models import (
    Category, VideoLesson, Test, Question, QuestionTypeRule,
    UserTestResult, UserTestAnswer, UserVideoProgress, UserActivity,
    Bookmark, StudyStreak, VideoNote, VideoRating,
    VideoComment, VideoPlaylist, PlaylistVideo, FlashcardSet, Flashcard
)

FILL_TYPES = ('fill_blank', 'summary_completion', 'notes_completion', 'sentence_completion', 
              'table_completion', 'short_answer')
MATCHING_TYPES = ('matching_headings', 'matching_features', 'matching_info', 
                  'matching_sentences', 'classification')
QUESTION_TYPE_LABELS = dict(Question.QUESTION_TYPES)


def _get_fill_blank_count(question):
    """Bo'sh joylar soni — model bilan bir xil (qisqa javob jadvalsiz qatorlar ham)."""
    return question.fill_blanks_count()


def _build_inline_fill_parts(question, ans_fields):
    """question_text ichidagi [1], [2], [3 4]... ni inline input. [3 4] — bitta yacheyka, label [3 4] ko'rsatiladi."""
    if not ans_fields:
        return None
    text = question.question_text or ''
    parts = []
    last = 0
    placed_nums = set()
    # Qavs ichidagi ixtiyoriy matn: [1], [2], [3 4] (2 ta so'z uchun bitta yacheyka)
    for idx, m in enumerate(re.finditer(r'\[([^\]]+)\]', text)):
        num = idx + 1
        if num > len(ans_fields):
            break
        label_raw = m.group(1).strip()
        if m.start() > last:
            parts.append({'type': 'text', 'content': text[last:m.start()]})
        field = next((f for f in ans_fields if f['num'] == num), None)
        if field:
            parts.append({'type': 'input', 'num': num, 'value': field.get('value', ''), 'label': label_raw})
            placed_nums.add(num)
        last = m.end()
    if last < len(text):
        parts.append({'type': 'text', 'content': text[last:]})
    for f in ans_fields:
        num = f.get('num')
        if num and num not in placed_nums:
            parts.append({'type': 'input', 'num': num, 'value': f.get('value', ''), 'label': str(num)})
    return parts if parts else None


FILL_MULTI_DISPLAY_TYPES = (
    'sentence_completion', 'table_completion', 'summary_completion',
    'notes_completion', 'fill_blank', 'short_answer',
)


def _card_fill_input_count(card):
    """Bitta savol kartochkasida nechta matn bo'sh joyi."""
    if card.get('inline_fill_parts'):
        return sum(1 for p in card['inline_fill_parts'] if p.get('type') == 'input')
    af = card.get('answer_fields') or []
    return len(af) if af else 1


def _get_question_context_extra(question, current_answer):
    """Savol turiga qarab answer_fields, matching_fields, list_options qaytarish"""
    ans_fields, matching_fields, list_options = [], [], []
    opts = question.options_json or {}
    correct = question.correct_answer_json or {}
    
    if question.question_type in FILL_TYPES:
        opts_fill = question.options_json or {}
        sa_rows = (
            opts_fill.get('short_answer_items') or []
            if question.question_type == 'short_answer'
            else []
        )
        if question.question_type == 'short_answer' and isinstance(sa_rows, list) and sa_rows:
            n_blanks = len(sa_rows)
            cl = question.get_correct_answers_list()
            if len(cl) < n_blanks:
                cl = list(cl) + [''] * (n_blanks - len(cl))
            ca_blanks = []
            if current_answer:
                try:
                    d = json.loads(current_answer)
                    ca_blanks = d if isinstance(d, list) else [str(d.get(str(i + 1), '')) for i in range(len(cl))]
                except (json.JSONDecodeError, TypeError):
                    ca_blanks = [x.strip() for x in str(current_answer).split(',')]
            while len(ca_blanks) < n_blanks:
                ca_blanks.append('')
            ans_fields = []
            for i, row in enumerate(sa_rows):
                if not isinstance(row, dict):
                    row = {}
                pr = (row.get('prompt') or row.get('text') or '').strip()
                try:
                    smw = int(row['max_words']) if row.get('max_words') not in (None, '') else None
                except (TypeError, ValueError):
                    smw = None
                if smw not in (1, 2, 3):
                    smw = question.get_max_words_per_blank()
                ans_fields.append({
                    'num': i + 1,
                    'value': ca_blanks[i] if i < len(ca_blanks) else '',
                    'prompt': pr,
                    'slot_max_words': smw,
                })
        else:
            n_blanks = _get_fill_blank_count(question)
            cl = question.get_correct_answers_list()
            if len(cl) < n_blanks:
                cl = list(cl) + [''] * (n_blanks - len(cl))
            ca_blanks = []
            if current_answer:
                if len(cl) > 1:
                    try:
                        d = json.loads(current_answer)
                        ca_blanks = d if isinstance(d, list) else [str(d.get(str(i+1), '')) for i in range(len(cl))]
                    except (json.JSONDecodeError, TypeError):
                        ca_blanks = [x.strip() for x in str(current_answer).split(',')]
                else:
                    ca_blanks = [current_answer]
            while len(ca_blanks) < len(cl):
                ca_blanks.append('')
            ans_fields = [{'num': i + 1, 'value': ca_blanks[i] if i < len(ca_blanks) else ''} for i in range(len(cl))]
    
    elif question.question_type in MATCHING_TYPES:
        items = opts.get('items', opts.get('paragraphs', []))
        if not items and isinstance(correct, dict):
            items = [{'num': k, 'label': f'Element {k}'} for k in sorted(correct.keys(), key=lambda x: int(x) if str(x).isdigit() else 0)]
        if question.question_type == 'matching_headings':
            options = opts.get('headings', []) or opts.get('options', [])
        else:
            options = opts.get('options', []) or opts.get('headings', [])
        if not options and isinstance(correct, dict):
            all_letters = sorted(set(str(v) for v in correct.values()))
            options = [{'letter': l, 'text': l} for l in all_letters]
        if isinstance(items, list) and (items or correct):
            cur_dict = {}
            if current_answer:
                try:
                    cur_dict = json.loads(current_answer) if isinstance(current_answer, str) and current_answer.startswith('{') else {}
                except json.JSONDecodeError:
                    pass
            opts_list = options if isinstance(options, list) else []
            for i, it in enumerate(items):
                num = it.get('num', i + 1) if isinstance(it, dict) else (i + 1)
                label = it.get('label', str(it)) if isinstance(it, dict) else str(it)
                val = cur_dict.get(str(num), '') if isinstance(cur_dict, dict) else ''
                matching_fields.append({
                    'num': num, 'label': label, 'value': val,
                    'options': [{'letter': o.get('letter', o) if isinstance(o, dict) else o, 'text': o.get('text', o) if isinstance(o, dict) else str(o)} for o in opts_list]
                })
    
    elif question.question_type == 'list_selection':
        options = opts.get('options', [])
        if not options:
            options = [{'letter': c, 'text': c} for c in (correct if isinstance(correct, list) else [])]
        cur_list = []
        if current_answer:
            try:
                cur_list = json.loads(current_answer) if isinstance(current_answer, str) and current_answer.startswith('[') else []
            except json.JSONDecodeError:
                cur_list = [current_answer]
        cur_set = set(str(x).lower() for x in cur_list)
        for o in options:
            letter = o.get('letter', o) if isinstance(o, dict) else o
            text = o.get('text', o) if isinstance(o, dict) else str(o)
            list_options.append({'letter': letter, 'text': text, 'checked': str(letter).lower() in cur_set})
    
    return ans_fields, matching_fields, list_options


@login_required
def dashboard(request):
    """Asosiy sahifa"""
    # Testlar uchun faqat yuqori darajadagi kategoriyalar (videodagi subkategoriyalar aralashmasin)
    categories = Category.objects.filter(is_active=True, show_on_site=True, parent__isnull=True).order_by('order', 'name')
    
    # Statistika (faqat interfeysda ko'rinadigan kategoriyalardagi testlar)
    total_videos = VideoLesson.objects.filter(is_active=True, category__show_on_site=True).count()
    total_tests = Test.objects.filter(is_active=True, category__show_on_site=True).count()
    
    # Foydalanuvchi statistikasi
    user_test_results = UserTestResult.objects.filter(user=request.user)
    user_video_progress = UserVideoProgress.objects.filter(user=request.user, watched=True)
    current_streak = StudyStreak.get_current_streak(request.user)
    
    user_stats = {
        'tests_completed': user_test_results.count(),
        'videos_watched': user_video_progress.count(),
        'average_score': user_test_results.aggregate(Avg('percentage'))['percentage__avg'] or 0,
        'current_streak': current_streak,
        'total_bookmarks': Bookmark.objects.filter(user=request.user).count(),
    }
    
    context = {
        'categories': categories,
        'total_videos': total_videos,
        'total_tests': total_tests,
        'user_stats': user_stats,
    }
    return render(request, 'core/dashboard.html', context)


@login_required
def video_list(request):
    """Video darslar ro'yxati"""
    category_slug = request.GET.get('category')
    subcategory_slug = request.GET.get('sub')
    search_query = request.GET.get('search', '')
    sort_by = request.GET.get('sort', 'order')
    
    videos = VideoLesson.objects.filter(is_active=True, category__show_on_site=True)
    
    selected_category = None
    selected_subcategory = None
    if category_slug:
        selected_category = Category.objects.filter(slug=category_slug, is_active=True).first()
        if subcategory_slug and selected_category:
            # Faqat shu otaning bolalaridan birini tanlash
            selected_subcategory = Category.objects.filter(
                slug=subcategory_slug, is_active=True, parent=selected_category
            ).first()
            if selected_subcategory:
                videos = videos.filter(category=selected_subcategory)
        elif selected_category:
            # Ota kategoriyadagi videolar ham, bolalar (subkategoriya) dagi videolar ham
            child_ids = list(
                Category.objects.filter(parent=selected_category, is_active=True).values_list('id', flat=True)
            )
            category_ids = [selected_category.id] + child_ids
            videos = videos.filter(category_id__in=category_ids)
    
    if search_query:
        videos = videos.filter(
            Q(title__icontains=search_query) | 
            Q(description__icontains=search_query) |
            Q(category__name__icontains=search_query)
        )
    
    # Sort
    if sort_by == 'newest':
        videos = videos.order_by('-created_at')
    elif sort_by == 'oldest':
        videos = videos.order_by('created_at')
    elif sort_by == 'title_asc':
        videos = videos.order_by('title')
    elif sort_by == 'title_desc':
        videos = videos.order_by('-title')
    elif sort_by == 'views':
        videos = videos.order_by('-views_count')
    else:  # order (default)
        videos = videos.order_by('order', 'created_at')
    
    videos = videos.select_related('category')
    
    # Foydalanuvchi progress va bookmarks
    user_progress = {}
    bookmarked_videos = set()
    if request.user.is_authenticated:
        progress_qs = UserVideoProgress.objects.filter(
            user=request.user,
            video__in=videos
        )
        for progress in progress_qs:
            user_progress[progress.video_id] = progress
        
        # Bookmarked videos
        bookmarks = Bookmark.objects.filter(user=request.user, video__in=videos)
        bookmarked_videos = {b.video_id for b in bookmarks}
    
    # Kategoriyalarga bo'lib ko'rsatish
    videos_by_category = {}
    page_obj = None
    # Yuqori darajadagi kategoriyalar (masalan: Grammar, Reading, Listening, Writing)
    top_categories = Category.objects.filter(
        is_active=True,
        show_on_site=True,
        parent__isnull=True,
    ).order_by('order', 'name')

    if not category_slug and not search_query:
        # Agar ota kategoriya tanlanmagan va qidiruv bo'lmasa, har bir top-level kategoriya uchun blok (ota + bolalar)
        for category in top_categories:
            child_ids = list(
                Category.objects.filter(parent=category, is_active=True).values_list('id', flat=True)
            )
            category_ids = [category.id] + child_ids
            qs = videos.filter(category_id__in=category_ids)
            if qs.exists():
                videos_by_category[category] = qs
    else:
        # Agar kategoriya tanlangan yoki qidiruv bo'lsa, oddiy ro'yxat
        paginator = Paginator(videos, 12)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        videos_by_category = None
    
    context = {
        'videos': page_obj if (category_slug or search_query) else None,
        'videos_by_category': videos_by_category,
        'categories': top_categories,
        'selected_category': category_slug,
        'selected_subcategory': subcategory_slug,
        'search_query': search_query,
        'sort_by': sort_by,
        'user_progress': user_progress,
        'bookmarked_videos': bookmarked_videos,
    }
    
    # HTMX request bo'lsa faqat video list qaytarish
    if request.headers.get('HX-Request'):
        return render(request, 'core/videos/partial_list.html', context)
    
    return render(request, 'core/videos/list.html', context)


@login_required
def video_detail(request, pk):
    """Video dars detallari"""
    video = get_object_or_404(VideoLesson, pk=pk, is_active=True)
    
    # YouTube ID ni ta'minlash - agar bo'lmasa, URL dan extract qilish
    if not video.youtube_id or len(video.youtube_id) != 11:
        if video.youtube_url:
            extracted_id = video.extract_youtube_id(video.youtube_url)
            if extracted_id and len(extracted_id) == 11:
                video.youtube_id = extracted_id
                video.save(update_fields=['youtube_id'])
    
    # Progress yaratish yoki olish
    progress, created = UserVideoProgress.objects.get_or_create(
        user=request.user,
        video=video
    )
    
    # Video boshlanganda watched=True
    if not progress.watched:
        progress.mark_as_watched()
        # Faollik yozish
        UserActivity.objects.create(
            user=request.user,
            activity_type='video_watch',
            related_object_id=video.pk,
            related_object_type='VideoLesson',
            metadata={'video_title': video.title}
        )
        # Study streak yangilash
        StudyStreak.update_streak(request.user)
    
    # Related videos
    related_videos = VideoLesson.objects.filter(
        category=video.category,
        is_active=True
    ).exclude(pk=video.pk)[:6]
    
    # Video notes
    video_notes = VideoNote.objects.filter(
        user=request.user,
        video=video
    ).order_by('timestamp')
    
    # User rating
    user_rating = VideoRating.objects.filter(
        user=request.user,
        video=video
    ).first()
    
    # Bookmark holati
    is_bookmarked = Bookmark.objects.filter(
        user=request.user,
        video=video
    ).exists()
    
    # Video comments (parent comments - javobsiz izohlar)
    video_comments = VideoComment.objects.filter(
        video=video,
        parent__isnull=True
    ).select_related('user').prefetch_related('replies__user').order_by('-created_at')[:20]
    
    # User playlists
    user_playlists = VideoPlaylist.objects.filter(user=request.user).order_by('-created_at')
    
    # Video qaysi playlistlarda bor?
    video_in_playlists = VideoPlaylist.objects.filter(
        user=request.user,
        videos=video
    ).values_list('id', flat=True)
    
    context = {
        'video': video,
        'progress': progress,
        'related_videos': related_videos,
        'video_notes': video_notes,
        'user_rating': user_rating,
        'is_bookmarked': is_bookmarked,
        'video_comments': video_comments,
        'user_playlists': user_playlists,
        'video_in_playlists': list(video_in_playlists),
    }
    return render(request, 'core/videos/detail.html', context)


@login_required
def test_list(request):
    """Testlar ro'yxati"""
    category_slug = request.GET.get('category')
    test_type = request.GET.get('type')
    difficulty = request.GET.get('difficulty')
    search_query = request.GET.get('search', '')
    sort_by = request.GET.get('sort', 'newest')
    
    tests = Test.objects.filter(is_active=True, category__show_on_site=True)
    
    if category_slug:
        tests = tests.filter(category__slug=category_slug)
    
    if test_type:
        tests = tests.filter(test_type=test_type)
    
    if difficulty:
        tests = tests.filter(difficulty=difficulty)
    
    if search_query:
        tests = tests.filter(
            Q(title__icontains=search_query) | 
            Q(description__icontains=search_query) |
            Q(category__name__icontains=search_query)
        )
    
    # Sort
    if sort_by == 'oldest':
        tests = tests.order_by('created_at')
    elif sort_by == 'title_asc':
        tests = tests.order_by('title')
    elif sort_by == 'title_desc':
        tests = tests.order_by('-title')
    elif sort_by == 'questions_asc':
        tests = tests.annotate(questions_count=Count('questions')).order_by('questions_count')
    elif sort_by == 'questions_desc':
        tests = tests.annotate(questions_count=Count('questions')).order_by('-questions_count')
    else:  # newest (default)
        tests = tests.order_by('-created_at')
    
    tests = tests.select_related('category').prefetch_related('questions')
    
    # Foydalanuvchi natijalari va bookmarks
    user_results = {}
    bookmarked_tests = set()
    if request.user.is_authenticated:
        results_qs = UserTestResult.objects.filter(
            user=request.user,
            test__in=tests
        ).select_related('test')
        for result in results_qs:
            user_results[result.test_id] = result
        
        # Bookmarked tests
        bookmarks = Bookmark.objects.filter(user=request.user, test__in=tests)
        bookmarked_tests = {b.test_id for b in bookmarks}
    
    # Testlar uchun faqat top-level kategoriyalar (videodagi subkategoriyalar aralashmasin)
    categories = Category.objects.filter(is_active=True, show_on_site=True, parent__isnull=True).order_by('order', 'name')
    test_types = Test.TEST_TYPES
    difficulty_levels = Test.DIFFICULTY_LEVELS
    
    # Adaptive Testing - yaxshilangan tavsiyalar
    recommended_tests = []
    weak_areas = []
    if request.user.is_authenticated:
        # Foydalanuvchining o'rtacha natijasini hisoblash
        user_avg_score = UserTestResult.objects.filter(
            user=request.user,
            completed_at__isnull=False
        ).aggregate(Avg('percentage'))['percentage__avg'] or 0
        
        # Kategoriya bo'yicha natijalar
        category_scores = UserTestResult.objects.filter(
            user=request.user,
            completed_at__isnull=False
        ).values('test__category__name', 'test__category__id').annotate(
            avg_score=Avg('percentage'),
            count=Count('id')
        ).order_by('avg_score')
        
        # Zaif tomonlarni aniqlash (50% dan past kategoriyalar)
        weak_category_ids = [item['test__category__id'] for item in category_scores if item['avg_score'] < 50 and item['count'] >= 2]
        weak_categories = Category.objects.filter(id__in=weak_category_ids).values('id', 'name', 'slug')
        weak_categories_dict = {cat['id']: cat for cat in weak_categories}
        
        weak_areas = [
            {
                'category_id': item['test__category__id'],
                'category_slug': weak_categories_dict.get(item['test__category__id'], {}).get('slug', ''),
                'category_name': item['test__category__name'],
                'avg_score': round(item['avg_score'], 1),
                'count': item['count']
            }
            for item in category_scores if item['avg_score'] < 50 and item['count'] >= 2
        ]
        
        # Foydalanuvchi darajasiga mos testlar
        if user_avg_score < 50:
            recommended_difficulty = 'easy'
        elif user_avg_score < 70:
            recommended_difficulty = 'medium'
        else:
            recommended_difficulty = 'hard'
        
        # Tavsiya qilingan testlar - zaif tomonlar bo'yicha
        recommended_query = Test.objects.filter(is_active=True, category__show_on_site=True)
        
        # Agar zaif tomonlar bo'lsa, ularni ustuvor qilish
        if weak_areas:
            weak_category_ids = [area['category_id'] for area in weak_areas]
            recommended_query = recommended_query.filter(
                Q(category_id__in=weak_category_ids, difficulty=recommended_difficulty) |
                Q(difficulty=recommended_difficulty)
            ).annotate(
                priority=Case(
                    When(category_id__in=weak_category_ids, then=1),
                    default=2,
                    output_field=IntegerField()
                )
            ).order_by('priority', 'created_at')
        else:
            recommended_query = recommended_query.filter(difficulty=recommended_difficulty).order_by('-created_at')
        
        # Foydalanuvchi ishlagan testlarni olib tashlash
        completed_test_ids = UserTestResult.objects.filter(
            user=request.user,
            completed_at__isnull=False
        ).values_list('test_id', flat=True).distinct()
        
        recommended_tests = recommended_query.exclude(
            id__in=completed_test_ids
        ).exclude(
            id__in=tests.values_list('id', flat=True)
        ).select_related('category')[:6]
    
    context = {
        'tests': tests,
        'categories': categories,
        'test_types': test_types,
        'difficulty_levels': difficulty_levels,
        'selected_category': category_slug,
        'selected_type': test_type,
        'selected_difficulty': difficulty,
        'search_query': search_query,
        'sort_by': sort_by,
        'user_results': user_results,
        'bookmarked_tests': bookmarked_tests,
        'recommended_tests': recommended_tests,
        'weak_areas': weak_areas,
        'user_avg_score': round(user_avg_score, 1) if request.user.is_authenticated else 0,
    }
    
    # HTMX request bo'lsa faqat test list qaytarish
    if request.headers.get('HX-Request'):
        return render(request, 'core/tests/partial_list.html', context)
    
    return render(request, 'core/tests/list.html', context)


@login_required
def test_collection_by_type(request, test_type):
    """Engnovate uslubida test collection sahifasi (reading/listening/writing)."""
    valid_types = {code for code, _ in Test.TEST_TYPES}
    if test_type not in valid_types:
        return redirect('core:test_list')

    search_query = request.GET.get('search', '').strip()
    question_type = request.GET.get('qtype', '').strip()
    length_filter = request.GET.get('length', '').strip()  # full | parts
    module_filter = request.GET.get('module', '').strip()  # academic | general

    tests = (
        Test.objects.filter(is_active=True, test_type=test_type, category__show_on_site=True)
        .select_related('category')
        .prefetch_related('questions')
        .annotate(questions_count=Count('questions', distinct=True))
    )

    if search_query:
        tests = tests.filter(
            Q(title__icontains=search_query)
            | Q(description__icontains=search_query)
            | Q(category__name__icontains=search_query)
        )

    if question_type:
        tests = tests.filter(questions__question_type=question_type).distinct()

    if length_filter == 'full':
        tests = tests.filter(questions_count__gte=30)
    elif length_filter == 'parts':
        tests = tests.filter(questions_count__lt=30)

    if module_filter == 'academic':
        tests = tests.filter(
            Q(title__icontains='academic') | Q(category__name__icontains='academic')
        )
    elif module_filter == 'general':
        tests = tests.filter(
            Q(title__icontains='general') | Q(category__name__icontains='general')
        )

    tests = tests.order_by('-created_at')

    available_types_qs = (
        Question.objects.filter(test__is_active=True, test__test_type=test_type)
        .values_list('question_type', flat=True)
        .distinct()
    )
    available_question_types = [
        (q_code, QUESTION_TYPE_LABELS.get(q_code, q_code.replace('_', ' ').title()))
        for q_code in available_types_qs
    ]
    available_question_types.sort(key=lambda item: item[1])

    paginator = Paginator(tests, 12)
    page_obj = paginator.get_page(request.GET.get('page'))

    for test in page_obj:
        title_lower = (test.title or '').lower()
        test.is_part_test = ('questions' in title_lower) or (getattr(test, 'questions_count', 0) < 30)
        first_question = test.questions.all().order_by('order').first()
        if first_question:
            test.primary_question_type = first_question.question_type
            test.primary_question_type_label = QUESTION_TYPE_LABELS.get(
                first_question.question_type,
                first_question.question_type.replace('_', ' ').title()
            )
        else:
            test.primary_question_type = ''
            test.primary_question_type_label = 'General'

    query_params = request.GET.copy()
    query_params.pop('page', None)

    context = {
        'collection_test_type': test_type,
        'collection_test_type_display': dict(Test.TEST_TYPES).get(test_type, test_type.title()),
        'tests': page_obj,
        'available_question_types': available_question_types,
        'selected_question_type': question_type,
        'selected_length': length_filter,
        'selected_module': module_filter,
        'search_query': search_query,
        'query_string': query_params.urlencode(),
    }
    return render(request, 'core/tests/type_collection.html', context)


@login_required
def test_detail(request, pk):
    """Test detallari va boshlash"""
    test = get_object_or_404(Test.objects.select_related('category'), pk=pk, is_active=True)
    if not getattr(test.category, 'show_on_site', True):
        return redirect('core:test_list')
    
    # Savollar ro'yxati
    questions = test.questions.all().order_by('order')
    
    # Foydalanuvchi oldingi natijalari
    previous_results = UserTestResult.objects.filter(
        user=request.user,
        test=test,
        completed_at__isnull=False
    ).order_by('-completed_at')
    
    previous_result = previous_results.first() if previous_results.exists() else None
    
    # Urinishlar soni
    attempts_count = previous_results.count()
    
    # Qayta ishlash mumkinmi?
    can_retake = test.allow_retake
    if test.max_attempts and attempts_count >= test.max_attempts:
        can_retake = False
    
    # Joriy test ishlanmoqdami?
    active_test = UserTestResult.objects.filter(
        user=request.user,
        test=test,
        completed_at__isnull=True
    ).order_by('-started_at').first()
    
    context = {
        'test': test,
        'questions': questions,
        'previous_result': previous_result,
        'previous_results': previous_results[:5],  # Oxirgi 5 ta natija
        'attempts_count': attempts_count,
        'can_retake': can_retake,
        'active_test': active_test,
    }
    return render(request, 'core/tests/detail.html', context)


@login_required
def test_take(request, pk):
    """Test ishlash"""
    test = get_object_or_404(Test.objects.select_related('category'), pk=pk, is_active=True)
    if not getattr(test.category, 'show_on_site', True):
        return redirect('core:test_list')
    
    # Test natijasi yaratish
    test_result, created = UserTestResult.objects.get_or_create(
        user=request.user,
        test=test,
        completed_at__isnull=True
    )
    
    if created:
        test_result.total_questions = test.total_questions
        # Urinish raqamini aniqlash
        previous_attempts = UserTestResult.objects.filter(
            user=request.user,
            test=test,
            completed_at__isnull=False
        ).count()
        test_result.attempt_number = previous_attempts + 1
        # Timer boshlash
        if test.duration_minutes:
            test_result.timer_started_at = timezone.now()
            test_result.timer_seconds_left = test.duration_minutes * 60
        test_result.save()
        
        # Faollik yozish
        UserActivity.objects.create(
            user=request.user,
            activity_type='test_start',
            related_object_id=test.pk,
            related_object_type='Test',
            metadata={'test_title': test.title}
        )
        # Study streak yangilash
        StudyStreak.update_streak(request.user)
    else:
        # Agar test to'xtatilgan bo'lsa, davom ettirish
        if test_result.is_paused:
            test_result.resume_test()
    
    # Javoblar
    answers = {}
    if test_result.answers_json:
        answers = test_result.answers_json
    
    # Hamma savollarni bitta sahifada ko'rsatamiz (2 variantli: avval Variant 1, keyin Variant 2)
    if getattr(test, 'variants_to_select', 1) == 2:
        questions = list(test.questions.all().order_by(Coalesce(F('variant'), Value(1)), 'order'))
    else:
        questions = list(test.questions.all().order_by('order'))
    total_questions = len(questions)

    if total_questions == 0:
        messages.warning(request, "Bu testda hali savollar qo'shilmagan. Admin orqali savollar qo'shing.")
        return redirect('core:test_detail', pk=test.pk)

    # Barcha javoblarni bir POST bilan saqlash
    if request.method == 'POST':
        updated_answers = {}
        single_choice = ('mcq', 'true_false', 'true_false_not_given', 'yes_no_not_given')
        fill_types = ('fill_blank', 'summary_completion', 'notes_completion', 'sentence_completion', 'table_completion', 'short_answer')

        for q in questions:
            val = ''
            if q.question_type in single_choice:
                if getattr(q, 'max_choices', 1) == 2:
                    selected = []
                    opts_json = q.options_json or {}
                    mcq_opts = opts_json.get('options') or []
                    if mcq_opts:
                        letters = [
                            str(o.get('letter', '')).strip().lower()
                            for o in mcq_opts
                            if o.get('letter')
                        ]
                    else:
                        letters = ['a', 'b', 'c', 'd']
                    for letter in letters:
                        if letter and request.POST.get(f'answer_{q.pk}_{letter}'):
                            selected.append(letter)
                    if selected:
                        val = json.dumps(sorted(selected))
                else:
                    val = (request.POST.get(f'answer_{q.pk}') or '').strip()
            elif q.question_type in fill_types:
                expected = _get_fill_blank_count(q)
                vals = []
                for i in range(1, expected + 1):
                    vals.append((request.POST.get(f'answer_{q.pk}_{i}') or '').strip())
                    #models to play
                if vals and not any(vals[1:]) and vals[0] and ',' in vals[0]:
                    vals = [v.strip() for v in vals[0].split(',')]
                if any(v for v in vals):
                    val = json.dumps(vals)
            elif q.question_type in MATCHING_TYPES:
                match_dict = {}
                opts = q.options_json or {}
                items = opts.get('items', [])
                if not items:
                    items = [{'num': i + 1} for i in range(len((q.correct_answer_json or {})))]
                for idx, it in enumerate(items):
                    num = str(it.get('num', idx + 1))
                    mval = (request.POST.get(f'match_{q.pk}_{num}') or '').strip()
                    if mval:
                        match_dict[num] = mval
                if match_dict:
                    val = json.dumps(match_dict)
            elif q.question_type == 'list_selection':
                selected = []
                for opt in (q.options_json or {}).get('options', []):
                    letter = str(opt.get('letter', '')).strip()
                    if letter and request.POST.get(f'list_{q.pk}_{letter}'):
                        selected.append(letter)
                if selected:
                    val = json.dumps(sorted(selected))
            elif q.question_type == 'essay':
                val = (request.POST.get(f'answer_{q.pk}') or '').strip()
            else:
                val = (request.POST.get(f'answer_{q.pk}') or '').strip()

            if val:
                updated_answers[str(q.pk)] = val

        test_result.answers_json = updated_answers
        test_result.save(update_fields=['answers_json'])
        answers = updated_answers

        if request.POST.get('finish_test') == '1':
            with transaction.atomic():
                total_slots_target = sum(
                    q.gradable_answer_slots() for q in questions if q.question_type != 'essay'
                )
                correct_count = 0
                for q in questions:
                    user_answer = (answers.get(str(q.pk), '') or '').strip()
                    if q.question_type == 'essay':
                        if user_answer:
                            UserTestAnswer.objects.update_or_create(
                                test_result=test_result,
                                question=q,
                                defaults={'user_answer': user_answer, 'is_correct': False},
                            )
                        continue
                    if q.mcq_dual_question_slots_enabled():
                        pts, _ = q.score_mcq_choose_two_dual(user_answer)
                        correct_count += pts
                        UserTestAnswer.objects.update_or_create(
                            test_result=test_result,
                            question=q,
                            defaults={
                                'user_answer': user_answer,
                                'is_correct': pts >= 2,
                            },
                        )
                        continue
                    if user_answer:
                        is_correct = q.check_user_answer(user_answer)
                        if is_correct:
                            correct_count += 1
                        UserTestAnswer.objects.update_or_create(
                            test_result=test_result,
                            question=q,
                            defaults={'user_answer': user_answer, 'is_correct': is_correct},
                        )

                test_result.total_questions = total_slots_target or total_questions
                test_result.correct_answers = correct_count
                test_result.wrong_answers = max(0, test_result.total_questions - correct_count)
                test_result.completed_at = timezone.now()
                test_result.attempt_number = test_result.attempt_number or 1
                test_result.time_taken = test_result.get_elapsed_time()
                test_result.calculate_score()
                test_result.refresh_from_db()

                UserActivity.objects.create(
                    user=request.user,
                    activity_type='test_complete',
                    related_object_id=test.pk,
                    related_object_type='Test',
                    metadata={'test_title': test.title, 'score': correct_count, 'total': total_questions}
                )
                StudyStreak.update_streak(request.user)

            return redirect('core:test_result', pk=test_result.pk)
        else:
            messages.success(request, "Javoblar saqlandi.")
            return redirect(request.path)

    answered_questions = [int(q_id) for q_id in answers.keys() if str(q_id).isdigit()]

    total_answer_slots = sum(
        q.gradable_answer_slots() for q in questions if q.question_type != 'essay'
    )
    answered_answer_slots = 0
    for q in questions:
        if q.question_type == 'essay':
            continue
        raw = answers.get(str(q.pk), '')
        if not raw or not str(raw).strip():
            continue
        if q.mcq_dual_question_slots_enabled():
            try:
                arr = json.loads(raw) if str(raw).strip().startswith('[') else []
                arr = arr if isinstance(arr, list) else []
                answered_answer_slots += min(len([x for x in arr if str(x).strip()]), 2)
            except (json.JSONDecodeError, TypeError):
                pass
        elif q.question_type in FILL_TYPES:
            try:
                vals = json.loads(raw) if str(raw).strip().startswith('[') else [raw]
                if isinstance(vals, list):
                    slots = q.gradable_answer_slots()
                    filled = sum(1 for v in vals if str(v).strip())
                    if slots > 1:
                        answered_answer_slots += min(filled, slots)
                    else:
                        answered_answer_slots += 1 if filled else 0
                elif str(raw).strip():
                    answered_answer_slots += 1
            except (json.JSONDecodeError, TypeError):
                answered_answer_slots += 1
        elif q.question_type in MATCHING_TYPES:
            try:
                d = json.loads(raw) if str(raw).strip().startswith('{') else {}
                if isinstance(d, dict):
                    slots = q.gradable_answer_slots()
                    filled = sum(1 for v in d.values() if str(v).strip())
                    if slots > 1:
                        answered_answer_slots += min(filled, slots)
                    else:
                        answered_answer_slots += 1 if filled else 0
                elif str(raw).strip():
                    answered_answer_slots += 1
            except (json.JSONDecodeError, TypeError):
                answered_answer_slots += 1 if str(raw).strip() else 0
        elif q.question_type == 'list_selection':
            try:
                arr = json.loads(raw) if str(raw).strip().startswith('[') else []
                if isinstance(arr, list) and len(arr) > 0:
                    slots = q.gradable_answer_slots()
                    answered_answer_slots += min(len([x for x in arr if str(x).strip()]), slots if slots else 1)
            except (json.JSONDecodeError, TypeError):
                pass
        else:
            answered_answer_slots += 1

    mcq_dual_question_pks = [q.pk for q in questions if q.mcq_dual_question_slots_enabled()]

    # Timer va vaqt ma'lumotlari
    timer_seconds_left = None
    timer_minutes = None
    timer_seconds = None
    if test.duration_minutes:
        timer_seconds_left = test_result.get_timer_seconds_left()
        # Timer ma'lumotlarini yangilash
        if timer_seconds_left is not None:
            test_result.timer_seconds_left = timer_seconds_left
            test_result.save(update_fields=['timer_seconds_left'])
            # Minutes va seconds ga ajratish
            timer_minutes = int(timer_seconds_left // 60)
            timer_seconds = int(timer_seconds_left % 60)
    
    elapsed_time = test_result.get_elapsed_time()
    answered_count = len(answered_questions)
    progress_percentage = (
        int((answered_answer_slots / max(total_answer_slots, 1)) * 100)
        if total_answer_slots > 0
        else 0
    )

    question_cards = []
    single_choice = ('mcq', 'true_false', 'true_false_not_given', 'yes_no_not_given')
    for q in questions:
        current_answer_val = answers.get(str(q.pk), '')
        ans_fields, match_flds, list_opts = _get_question_context_extra(q, current_answer_val)
        mcq_opts = []
        mcq_single_banner = ''
        mcq_choose_two_banner = ''
        fill_banner_text = ''
        if q.question_type in single_choice:
            # True/False va Yes/No/Not Given da har doim standart variantlar (admin qanday yozganidan qat'iy nazar)
            if q.question_type == 'true_false':
                mcq_opts = [{'letter': 'a', 'text': 'TRUE'}, {'letter': 'b', 'text': 'FALSE'}]
            elif q.question_type == 'true_false_not_given':
                mcq_opts = [{'letter': 'a', 'text': 'TRUE'}, {'letter': 'b', 'text': 'FALSE'}, {'letter': 'c', 'text': 'NOT GIVEN'}]
            elif q.question_type == 'yes_no_not_given':
                mcq_opts = [{'letter': 'a', 'text': 'YES'}, {'letter': 'b', 'text': 'NO'}, {'letter': 'c', 'text': 'NOT GIVEN'}]
            else:
                opts_json = q.options_json or {}
                mcq_opts = opts_json.get('options') or []
                if not mcq_opts and any([q.option_a, q.option_b, q.option_c, q.option_d]):
                    mcq_opts = []
                    for letter, txt in [('a', q.option_a), ('b', q.option_b), ('c', q.option_c), ('d', q.option_d)]:
                        if txt:
                            mcq_opts.append({'letter': letter, 'text': txt})
        inline_parts = None
        if q.question_type in FILL_TYPES and ans_fields and re.search(r'\[\d+\]', q.question_text or ''):
            inline_parts = _build_inline_fill_parts(q, ans_fields)
        task_images = []
        if test.test_type == 'writing' and hasattr(q, 'get_task_images'):
            task_images = q.get_task_images(request)
        mcq_choose_two = (q.question_type in single_choice and getattr(q, 'max_choices', 1) == 2)
        if mcq_choose_two and q.question_type == 'mcq':
            # max_choices=2 bo'lsa (Choose TWO...) banner matnini chiqaramiz.
            letters = [
                str(o.get('letter', '')).strip().upper()
                for o in (mcq_opts or [])
                if isinstance(o, dict) and o.get('letter')
            ]
            # fallback: option_a/b/c/d bor bo'lsa
            if not letters and any([q.option_a, q.option_b, q.option_c, q.option_d]):
                letters = [x for x in ['A', 'B', 'C', 'D'] if (x == 'A' and q.option_a) or (x == 'B' and q.option_b) or (x == 'C' and q.option_c) or (x == 'D' and q.option_d)]

            letters = [x for x in letters if x]
            uniq = []
            for l in letters:
                if l not in uniq:
                    uniq.append(l)
            letters = uniq
            # A-E kabi ketma-ket bo'lsa diapazon, aks holda ro'yxat.
            if len(letters) >= 2:
                # sort by alphabet order
                letters_sorted = sorted(letters, key=lambda x: ord(x[0]) if x else 0)
                first = letters_sorted[0][0]
                last = letters_sorted[-1][0]
                is_consecutive = all(
                    (ord(letters_sorted[i][0]) - ord(first)) == i + (ord(letters_sorted[i][0]) - ord(first) - i)
                    for i in range(len(letters_sorted))
                )
                if first.isalpha() and last.isalpha():
                    # oddiy consecutive check: A..E
                    ords = [ord(x[0]) for x in letters_sorted if x and x[0].isalpha()]
                    is_consecutive = bool(ords) and ords == list(range(min(ords), max(ords) + 1))
                    if is_consecutive:
                        letters_label = f"{first}–{last}"
                    else:
                        letters_label = ', '.join(letters_sorted)
                else:
                    letters_label = ', '.join(letters_sorted)

                mcq_choose_two_banner = f"Choose TWO letters, {letters_label}. Write the correct letters in boxes on your answer sheet."
        if q.question_type == 'mcq' and not mcq_choose_two and mcq_opts:
            lets = [
                str(o.get('letter', '')).strip().upper()
                for o in mcq_opts
                if isinstance(o, dict) and o.get('letter')
            ]
            lets = [x for x in lets if x]
            if len(lets) >= 2:
                mcq_single_banner = (
                    'Choose the correct letter, '
                    + ', '.join(lets[:-1])
                    + ' or '
                    + lets[-1]
                    + '. Write the correct letter in boxes on your answer sheet.'
                )
            elif len(lets) == 1:
                mcq_single_banner = (
                    f'Choose the correct letter, {lets[0]}. Write the answer in the box on your answer sheet.'
                )
        current_answer_list = []
        if mcq_choose_two and current_answer_val:
            try:
                raw = current_answer_val.strip()
                if raw.startswith('['):
                    current_answer_list = [str(x).strip().lower() for x in json.loads(raw) if x]
                else:
                    current_answer_list = [raw.lower()] if raw else []
            except (TypeError, json.JSONDecodeError):
                current_answer_list = [current_answer_val.lower()] if current_answer_val else []
        max_w = q.get_max_words_per_blank()
        # IELTS uslubidagi banner matni (Engnovate’ga yaqin)
        if q.question_type in FILL_TYPES:
            inst = ''
            try:
                inst = (q.options_json or {}).get('instruction') or ''
            except Exception:
                inst = ''
            qt = q.question_text or ''
            u = (inst + "\n" + qt).upper()
            if 'ONE WORD AND/OR A NUMBER' in u or 'ONE WORD AND/OR A NUMBER' in inst.upper():
                fill_banner_text = 'Write ONE WORD AND/OR A NUMBER for each answer.'
            elif 'ONE WORD ONLY' in u or 'ONE WORD ONLY' in inst.upper():
                fill_banner_text = 'Write ONE WORD ONLY for each answer.'
            elif 'NO MORE THAN TWO WORDS' in u or 'TWO WORDS' in u:
                fill_banner_text = 'Choose NO MORE THAN TWO WORDS from the passage for each answer.'
            elif 'NO MORE THAN THREE WORDS' in u or 'THREE WORDS' in u:
                fill_banner_text = 'Choose NO MORE THAN THREE WORDS from the passage for each answer.'
            else:
                # fallback (max_w bo‘yicha)
                if max_w == 1:
                    fill_banner_text = 'Choose ONE WORD from the passage for each answer.'
                elif max_w == 2:
                    fill_banner_text = 'Choose NO MORE THAN TWO WORDS from the passage for each answer.'
                elif max_w == 3:
                    fill_banner_text = 'Choose NO MORE THAN THREE WORDS from the passage for each answer.'
        matching_ref_opts = []
        if match_flds and match_flds[0].get('options'):
            matching_ref_opts = list(match_flds[0]['options'])
        sa_list = (q.options_json or {}).get('short_answer_items') or []
        sa_standalone = (
            q.question_type == 'short_answer'
            and isinstance(sa_list, list)
            and len(sa_list) > 0
        )
        slot_mws = [f.get('slot_max_words') for f in ans_fields if isinstance(f, dict) and f.get('slot_max_words')]
        fill_slots_same = sa_standalone and slot_mws and len(set(slot_mws)) == 1
        question_cards.append({
            'question': q,
            'current_answer': current_answer_val,
            'current_answer_list': current_answer_list,
            'mcq_choose_two': mcq_choose_two,
            'answer_fields': ans_fields,
            'matching_fields': match_flds,
            'matching_ref_options': matching_ref_opts,
            'list_options': list_opts,
            'mcq_options': mcq_opts,
            'inline_fill_parts': inline_parts,
            'question_images': task_images,
            'fill_max_words': max_w if not fill_slots_same else slot_mws[0],
            'sa_standalone_rows': sa_standalone,
            'fill_mixed_word_limits': sa_standalone and slot_mws and len(set(slot_mws)) > 1,
            'mcq_single_banner': mcq_single_banner,
            'fill_banner_text': fill_banner_text,
            'short_answer_banner_text': fill_banner_text,
            'mcq_choose_two_banner': mcq_choose_two_banner,
        })

    # Savol raqamlari: MCQ 2 tanlov → 21,22; bir nechta bo'sh joy → 23–26; boshqa → bittadan
    running_display = 1
    for card in question_cards:
        q = card['question']
        card['fill_multi_slots'] = False
        card['matching_multi_slots'] = False
        if q.mcq_dual_question_slots_enabled():
            card['display_order'] = running_display
            card['display_order_2'] = running_display + 1
            card['mcq_dual_slots'] = True
            running_display += 2
        elif q.question_type in FILL_MULTI_DISPLAY_TYPES and _card_fill_input_count(card) > 1:
            k = _card_fill_input_count(card)
            nums = list(range(running_display, running_display + k))
            running_display += k
            card['display_order'] = nums[0]
            card['display_order_2'] = None
            card['display_order_end'] = nums[-1]
            card['mcq_dual_slots'] = False
            card['fill_multi_slots'] = True
            card['fill_global_nums'] = nums
            gi = 0
            if card.get('inline_fill_parts'):
                for p in card['inline_fill_parts']:
                    if p.get('type') == 'input':
                        p['global_num'] = nums[gi] if gi < len(nums) else nums[-1]
                        gi += 1
            for i, f in enumerate(card.get('answer_fields') or []):
                f['global_num'] = nums[i] if i < len(nums) else nums[-1]
        elif q.question_type in MATCHING_TYPES and len(card.get('matching_fields') or []) > 1:
            mfs = card['matching_fields']
            k = len(mfs)
            nums = list(range(running_display, running_display + k))
            running_display += k
            card['display_order'] = nums[0]
            card['display_order_2'] = None
            card['display_order_end'] = nums[-1]
            card['mcq_dual_slots'] = False
            card['matching_multi_slots'] = True
            card['matching_global_nums'] = nums
            for i, mf in enumerate(mfs):
                mf['global_num'] = nums[i]
        elif q.question_type in MATCHING_TYPES and len(card.get('matching_fields') or []) == 1:
            d = running_display
            running_display += 1
            card['display_order'] = d
            card['display_order_2'] = None
            card['mcq_dual_slots'] = False
            card['matching_fields'][0]['global_num'] = d
        else:
            d = running_display
            running_display += 1
            card['display_order'] = d
            card['display_order_2'] = None
            card['mcq_dual_slots'] = False
            if card.get('inline_fill_parts'):
                for p in card['inline_fill_parts']:
                    if p.get('type') == 'input':
                        p['global_num'] = d
            for f in card.get('answer_fields') or []:
                f['global_num'] = d

    # Part bloklari: 2 variantli testda part = variant (1 yoki 2); boshqalarida explicit part yoki default
    part_indexes = []
    explicit_parts = []
    if getattr(test, 'variants_to_select', 1) == 2:
        part_indexes = [getattr(q, 'variant', None) or 1 for q in questions]
    else:
        for idx, q in enumerate(questions):
            opts = q.options_json or {}
            raw_part = opts.get('part', opts.get('section'))
            if raw_part is None:
                explicit_parts = []
                break
            try:
                explicit_parts.append(int(raw_part))
            except (TypeError, ValueError):
                explicit_parts = []
                break

    if not part_indexes and explicit_parts and len(explicit_parts) == len(questions):
        min_part = min(explicit_parts) if explicit_parts else 1
        part_indexes = [max(1, p - min_part + 1) for p in explicit_parts]
    elif not part_indexes:
        # Part raqami belgilanmagan: Reading da passage + savollar soniga qarab (13 ta = 1 part)
        default_parts = 1
        if test.test_type == 'reading':
            passages_list_for_count = test.get_reading_passages()
            passage_count = len(passages_list_for_count) if passages_list_for_count else 0
            # Savollar soni bo'yicha: 1–13 → 1 part, 14–26 → 2 part, 27+ → 3 part (IELTS uslubi)
            if total_questions <= 13:
                parts_by_questions = 1
            elif total_questions <= 26:
                parts_by_questions = 2
            else:
                parts_by_questions = 3
            default_parts = max(1, min(passage_count, parts_by_questions))
        elif test.test_type == 'listening':
            # Har partda 10 ta savol: Part 1 = 1–10, Part 2 = 11–20, Part 3 = 21–30, Part 4 = 31–40
            default_parts = min(4, max(1, ((total_questions or 0) + 9) // 10))
        elif test.test_type == 'writing':
            default_parts = 2
        if test.test_type != 'listening':
            default_parts = max(1, min(default_parts, total_questions or 1))
        #mmm

        if test.test_type == 'reading' and default_parts <= 3:
            # IELTS: Part 1 = 13 ta, Part 2 = 13 ta, Part 3 = qolgani (14 yoki kam)
            ranges = []
            if default_parts >= 1:
                ranges.append((0, min(13, total_questions)))
            if default_parts >= 2:
                ranges.append((13, min(26, total_questions)))
            if default_parts >= 3:
                ranges.append((26, total_questions))
        elif test.test_type == 'listening':
            # Listening: har partda 10 ta savol (1–10, 11–20, 21–30, 31–40)
            t = total_questions or 0
            ranges = []
            if t > 0:
                ranges.append((0, min(10, t)))
            if t > 10:
                ranges.append((10, min(20, t)))
            if t > 20:
                ranges.append((20, min(30, t)))
            if t > 30:
                ranges.append((30, t))
        else:
            base_size = total_questions // default_parts if default_parts else total_questions
            extra = total_questions % default_parts if default_parts else 0
            ranges = []
            start = 0
            for p in range(default_parts):
                size = base_size + (1 if p < extra else 0)
                end = start + size
                ranges.append((start, end))
                start = end

        for idx in range(total_questions):
            assigned = 1
            for p_idx, (s, e) in enumerate(ranges, start=1):
                if s <= idx < e:
                    assigned = p_idx
                    break
            part_indexes.append(assigned)

    # Reading: faqat 0 yoki 1 ta passage bo'lsa — barcha savollar bitta partda; 2 variantli testda part_indexes allaqachon variant bo'yicha
    if test.test_type == 'reading' and part_indexes and getattr(test, 'variants_to_select', 1) != 2:
        passage_count = len(test.get_reading_passages()) if hasattr(test, 'get_reading_passages') else 0
        if passage_count <= 1:
            part_indexes = [1] * len(part_indexes)

    part_groups = []
    for idx, card in enumerate(question_cards):
        part_no = part_indexes[idx] if idx < len(part_indexes) else 1
        if not part_groups or part_groups[-1]['part_number'] != part_no:
            part_groups.append({
                'part_number': part_no,
                'title': f'Part {part_no}',
                'cards': [],
                'start_order': card['question'].order,
                'end_order': card['question'].order,
            })
        part_groups[-1]['cards'].append(card)
        part_groups[-1]['end_order'] = card['question'].order

    type_shart_map = {r.question_type: (r.shart_text or '').strip() for r in QuestionTypeRule.objects.all()}
    for pg in part_groups:
        pg['question_count'] = len(pg['cards'])
        pg['slug'] = f"part-{pg['part_number']}"
        if getattr(test, 'variants_to_select', 1) == 2:
            pg['title'] = f"Variant {pg['part_number']}"
        elif test.test_type == 'writing':
            pg['title'] = f"Task {pg['part_number']}"
        blank_buttons = []
        # Reading va Listening testlarda dockda faqat savol raqamlari (blanklar bo'yicha emas)
        # Inline blanklar bo'yicha alohida navigatsiya faqat boshqa test turlarida kerak bo'lishi mumkin
        use_question_buttons_only = (test.test_type in ('reading', 'listening'))
        for card in pg['cards']:
            q = card['question']
            if not use_question_buttons_only and pg['part_number'] == 1 and (card.get('inline_fill_parts') or (card.get('answer_fields') and len(card['answer_fields']) > 1)):
                if card.get('inline_fill_parts'):
                    for p in card['inline_fill_parts']:
                        if p.get('type') == 'input':
                            blank_buttons.append({
                                'num': p['num'],
                                'blank_id': f"blank-{q.pk}-{p['num']}",
                                'card_anchor': str(p['num']),
                                'is_blank': True,
                            })
                else:
                    for f in card['answer_fields']:
                        blank_buttons.append({
                            'num': f['num'],
                            'blank_id': f"blank-{q.pk}-{f['num']}",
                            'card_anchor': str(f['num']),
                            'is_blank': True,
                        })
            else:
                do = card.get('display_order', q.order)
                if card.get('fill_multi_slots') and card.get('fill_global_nums'):
                    first = card['fill_global_nums'][0]
                    for gn in card['fill_global_nums']:
                        blank_buttons.append({
                            'num': gn,
                            'question_id': f'question-{first}',
                            'card_anchor': str(first),
                            'is_blank': False,
                        })
                elif card.get('matching_multi_slots') and card.get('matching_global_nums'):
                    first = card['matching_global_nums'][0]
                    for gn in card['matching_global_nums']:
                        blank_buttons.append({
                            'num': gn,
                            'question_id': f'question-{first}',
                            'card_anchor': str(first),
                            'is_blank': False,
                        })
                elif card.get('mcq_dual_slots'):
                    do2 = card.get('display_order_2') or (do + 1)
                    blank_buttons.append({
                        'num': do,
                        'question_id': f'question-{do}',
                        'card_anchor': str(do),
                        'is_blank': False,
                    })
                    blank_buttons.append({
                        'num': do2,
                        'question_id': f'question-{do}',
                        'card_anchor': str(do),
                        'is_blank': False,
                    })
                else:
                    blank_buttons.append({
                        'num': do,
                        'question_id': f'question-{do}',
                        'card_anchor': str(do),
                        'is_blank': False,
                    })
        # Reading: partda savollar bor lekin tugmalar bo'sh qolsa (eski ma'lumot), har savol uchun tugma yaratamiz
        if test.test_type == 'reading' and not blank_buttons and pg['cards']:
            for card in pg['cards']:
                q = card['question']
                do = card.get('display_order', q.order)
                if card.get('fill_multi_slots') and card.get('fill_global_nums'):
                    first = card['fill_global_nums'][0]
                    for gn in card['fill_global_nums']:
                        blank_buttons.append({'num': gn, 'question_id': f'question-{first}', 'card_anchor': str(first), 'is_blank': False})
                elif card.get('matching_multi_slots') and card.get('matching_global_nums'):
                    first = card['matching_global_nums'][0]
                    for gn in card['matching_global_nums']:
                        blank_buttons.append({'num': gn, 'question_id': f'question-{first}', 'card_anchor': str(first), 'is_blank': False})
                elif card.get('mcq_dual_slots'):
                    do2 = card.get('display_order_2') or (do + 1)
                    blank_buttons.append({'num': do, 'question_id': f'question-{do}', 'card_anchor': str(do), 'is_blank': False})
                    blank_buttons.append({'num': do2, 'question_id': f'question-{do}', 'card_anchor': str(do), 'is_blank': False})
                else:
                    blank_buttons.append({'num': do, 'question_id': f'question-{do}', 'card_anchor': str(do), 'is_blank': False})
        pg['blank_buttons'] = blank_buttons
        # range_label: har bir part o'z oralig'ini ko'rsatadi (Part 1: 1-13, Part 2: 14-26, Part 3: 27-40)
        if blank_buttons and all(b.get('is_blank') for b in blank_buttons):
            nums = [b['num'] for b in blank_buttons]
            pg['range_label'] = f"{min(nums)}-{max(nums)}" if len(nums) > 1 else str(nums[0])
        else:
            dock_nums = [b['num'] for b in blank_buttons if b.get('question_id')]
            if dock_nums:
                lo, hi = min(dock_nums), max(dock_nums)
                pg['range_label'] = f'{lo}-{hi}' if lo != hi else str(lo)
            else:
                pg['range_label'] = f"{pg['start_order']}-{pg['end_order']}" if pg['start_order'] != pg['end_order'] else str(pg['start_order'])

        # Savol turi bo'yicha guruhlash: har bir turdan oldin shart (QuestionTypeRule) ko'rsatiladi
        type_blocks = []
        for card in pg['cards']:
            q_type = card['question'].question_type
            d0 = card['display_order']
            d1 = card.get('display_order_end', d0)
            if type_blocks and type_blocks[-1]['question_type'] == q_type:
                type_blocks[-1]['cards'].append(card)
                type_blocks[-1]['end_order'] = max(type_blocks[-1]['end_order'], d1)
            else:
                type_blocks.append({
                    'question_type': q_type,
                    'shart_text': type_shart_map.get(q_type, ''),
                    'cards': [card],
                    'start_order': d0,
                    'end_order': d1,
                })
        if not type_blocks and pg['cards']:
            c0, c1 = pg['cards'][0], pg['cards'][-1]
            type_blocks = [{
                'question_type': '', 'shart_text': '', 'cards': pg['cards'],
                'start_order': c0['display_order'],
                'end_order': c1.get('display_order_end', c1['display_order']),
            }]
        pg['type_blocks'] = type_blocks

    # Reading: har bir partga mos passage (2 variantli da [v1_list, v2_list] — har partga o'sha variantning birinchi passage'i)
    if test.test_type == 'reading':
        passages_list = test.get_reading_passages()
        for idx, pg in enumerate(part_groups):
            if idx < len(passages_list):
                p = passages_list[idx]
                if isinstance(p, list):
                    pg['passage'] = p[0] if p else None
                else:
                    pg['passage'] = p
            else:
                pg['passage'] = None

    # Listening: har bir part uchun "Listen From Here" da boshlash vaqti (birinchi savolning audio_timestamp)
    if test.test_type == 'listening':
        for pg in part_groups:
            first_ts = None
            if pg.get('cards'):
                q = pg['cards'][0].get('question')
                if q and getattr(q, 'audio_timestamp', None) is not None:
                    first_ts = float(q.audio_timestamp)
            pg['audio_start_time'] = first_ts if first_ts is not None else 0

    current_question = questions[0] if questions else None
    # "Questions 1-10" yoki "Questions 1-7" ko'rsatish uchun
    first_pg = part_groups[0] if part_groups else {}
    first_blanks = first_pg.get('blank_buttons', [])
    if first_blanks and all(b.get('is_blank') for b in first_blanks):
        nums = [b['num'] for b in first_blanks]
        questions_range_display = f"{min(nums)}-{max(nums)}" if len(nums) > 1 else str(nums[0])
    else:
        questions_range_display = f"1-{total_questions}" if total_questions > 1 else "1"

    # Blank-based answered count (notes/summary: 3/10)
    total_blanks = len(first_blanks) if first_blanks and all(b.get('is_blank') for b in first_blanks) else 0
    answered_blanks = 0
    if total_blanks:
        for card in first_pg.get('cards', []):
            q = card.get('question')
            if not q:
                continue
            raw = answers.get(str(q.pk), '')
            if not raw:
                continue
            try:
                vals = json.loads(raw) if isinstance(raw, str) and raw.startswith('[') else [raw]
                answered_blanks += sum(1 for v in vals if str(v).strip())
            except (json.JSONDecodeError, TypeError):
                if str(raw).strip():
                    answered_blanks += 1

    context = {
        'test': test,
        'questions_range_display': questions_range_display,
        'total_blanks': total_blanks,
        'answered_blanks': answered_blanks,
        'test_result': test_result,
        'current_question': current_question,
        'current_answer': answers.get(str(current_question.pk), '') if current_question else '',
        'question_number': answered_count,
        'total_questions': total_questions,
        'total_answer_slots': total_answer_slots,
        'answered_answer_slots': answered_answer_slots,
        'mcq_dual_question_pks': mcq_dual_question_pks,
        'mcq_dual_question_pks_json': json.dumps(mcq_dual_question_pks),
        'progress_percentage': progress_percentage,
        'answered_questions': answered_questions,
        'question_cards': question_cards,
        'part_groups': part_groups,
        'timer_seconds_left': timer_seconds_left,
        'timer_minutes': timer_minutes,
        'timer_seconds': timer_seconds,
        'elapsed_time': elapsed_time,
        'is_paused': test_result.is_paused,
        'flashcard_sets': FlashcardSet.objects.filter(user=request.user).order_by('name'),
    }

    return render(request, 'core/tests/take.html', context)


@login_required
@require_POST
def add_test_flashcard(request, pk):
    """Test sahifasidan flashcard qo'shish."""
    test = get_object_or_404(Test.objects.select_related('category'), pk=pk, is_active=True)
    if not getattr(test.category, 'show_on_site', True):
        return JsonResponse({'success': False, 'error': 'Test mavjud emas.'}, status=404)

    term = (request.POST.get('term') or '').strip()
    definition = (request.POST.get('definition') or '').strip()
    set_id = (request.POST.get('set_id') or '').strip()
    new_set_name = (request.POST.get('new_set_name') or '').strip()
    question_id = (request.POST.get('question_id') or '').strip()

    if not term:
        return JsonResponse({'success': False, 'error': 'Term bo\'sh bo\'lishi mumkin emas.'}, status=400)

    flashcard_set = None
    if new_set_name:
        flashcard_set, _ = FlashcardSet.objects.get_or_create(
            user=request.user,
            name=new_set_name[:120],
        )
    elif set_id:
        flashcard_set = FlashcardSet.objects.filter(pk=set_id, user=request.user).first()
    if not flashcard_set:
        return JsonResponse({'success': False, 'error': 'Set tanlang yoki yangi set nomini kiriting.'}, status=400)

    source_question = None
    if question_id.isdigit():
        source_question = Question.objects.filter(pk=int(question_id), test=test).first()

    card = Flashcard.objects.create(
        user=request.user,
        flashcard_set=flashcard_set,
        term=term[:255],
        definition=definition[:2000],
        source_test=test,
        source_question=source_question,
    )
    return JsonResponse({
        'success': True,
        'id': card.pk,
        'set_id': flashcard_set.pk,
        'set_name': flashcard_set.name,
    })


@login_required
def test_retake(request, pk):
    """Testni qayta ishlash"""
    test = get_object_or_404(Test.objects.select_related('category'), pk=pk, is_active=True)
    if not getattr(test.category, 'show_on_site', True):
        return redirect('core:test_list')
    
    # Qayta ishlashga ruxsat bormi?
    if not test.allow_retake:
        messages.error(request, 'Bu testni qayta ishlashga ruxsat berilmagan.')
        if request.headers.get('HX-Request'):
            return redirect('core:test_detail', pk=test.pk)
        return redirect('core:test_detail', pk=test.pk)
    
    # Urinishlar sonini tekshirish
    attempts_count = UserTestResult.objects.filter(
        user=request.user,
        test=test,
        completed_at__isnull=False
    ).count()
    
    if test.max_attempts and attempts_count >= test.max_attempts:
        messages.error(request, f'Siz maksimal {test.max_attempts} marta urinish qildingiz.')
        if request.headers.get('HX-Request'):
            return redirect('core:test_detail', pk=test.pk)
        return redirect('core:test_detail', pk=test.pk)
    
    # Faol testni o'chirish (agar bo'lsa)
    UserTestResult.objects.filter(
        user=request.user,
        test=test,
        completed_at__isnull=True
    ).delete()
    
    # Yangi test boshlash
    if request.headers.get('HX-Request'):
        # HTMX request - test_take sahifasini yuklash
        return redirect('core:test_take', pk=test.pk)
    return redirect('core:test_take', pk=test.pk)


@login_required
def test_result(request, pk):
    """Test natijasi"""
    test_result = get_object_or_404(
        UserTestResult.objects.select_related('test', 'user').prefetch_related('test__questions', 'answers'),
        pk=pk, 
        user=request.user
    )
    
    # Qayta ishlash mumkinmi?
    can_retake = test_result.test.allow_retake
    if test_result.test.max_attempts:
        attempts_count = UserTestResult.objects.filter(
            user=request.user,
            test=test_result.test,
            completed_at__isnull=False
        ).count()
        if attempts_count >= test_result.test.max_attempts:
            can_retake = False
    
    # Agar test yakunlanmagan bo'lsa, yakunlash
    if not test_result.completed_at:
        # Transaction ichida barcha operatsiyalarni bajarish
        try:
            with transaction.atomic():
                # Javoblarni tekshirish va natijani hisoblash
                answers = test_result.answers_json
                correct = 0

                # Barcha savollarni bir martada olish
                questions = list(test_result.test.questions.all().order_by('order'))
                
                total_slots_target = sum(
                    q.gradable_answer_slots() for q in questions if q.question_type != 'essay'
                )
                for question in questions:
                    user_answer = (answers.get(str(question.pk), '') or '').strip()
                    if question.question_type == 'essay':
                        if user_answer:
                            UserTestAnswer.objects.update_or_create(
                                test_result=test_result,
                                question=question,
                                defaults={'user_answer': user_answer, 'is_correct': False},
                            )
                        continue
                    if question.mcq_dual_question_slots_enabled():
                        pts, _ = question.score_mcq_choose_two_dual(user_answer)
                        correct += pts
                        UserTestAnswer.objects.update_or_create(
                            test_result=test_result,
                            question=question,
                            defaults={
                                'user_answer': user_answer,
                                'is_correct': pts >= 2,
                            },
                        )
                        continue
                    if user_answer:
                        is_correct = question.check_user_answer(user_answer)
                        if is_correct:
                            correct += 1
                        UserTestAnswer.objects.update_or_create(
                            test_result=test_result,
                            question=question,
                            defaults={
                                'user_answer': user_answer,
                                'is_correct': is_correct,
                            },
                        )

                test_result.total_questions = total_slots_target or len(questions)
                test_result.correct_answers = correct
                test_result.wrong_answers = max(0, test_result.total_questions - correct)
                test_result.completed_at = timezone.now()
                test_result.attempt_number = test_result.attempt_number or 1
                # Time tracking - sarflangan vaqtni hisoblash
                test_result.time_taken = test_result.get_elapsed_time()
                test_result.calculate_score()
                test_result.save()
                
                # Faollik yozish
                UserActivity.objects.create(
                    user=request.user,
                    activity_type='test_complete',
                    related_object_id=test_result.test.pk,
                    related_object_type='Test',
                    metadata={
                        'test_title': test_result.test.title,
                        'score': test_result.score,
                        'percentage': test_result.percentage
                    }
                )
                # Study streak yangilash
                StudyStreak.update_streak(request.user)
        except Exception as e:
            messages.error(request, f'Xatolik yuz berdi: {str(e)}')
            return redirect('core:test_detail', pk=test_result.test.pk)
    
    # Javoblar ro'yxati
    user_answers = {}
    for answer in test_result.answers.all():
        user_answers[answer.question_id] = answer
    
    # Vaqtni formatlash
    time_taken_hours = None
    time_taken_minutes = None
    time_taken_seconds = None
    if test_result.time_taken:
        if test_result.time_taken >= 3600:
            time_taken_hours = test_result.time_taken // 3600
            remaining_seconds = test_result.time_taken % 3600
            time_taken_minutes = remaining_seconds // 60
        elif test_result.time_taken >= 60:
            time_taken_minutes = test_result.time_taken // 60
            time_taken_seconds = test_result.time_taken % 60
        else:
            time_taken_seconds = test_result.time_taken
    
    # Oldingi natijalar bilan solishtirish (Test Results Comparison)
    previous_results = UserTestResult.objects.filter(
        user=request.user,
        test=test_result.test,
        completed_at__isnull=False
    ).exclude(pk=test_result.pk).order_by('-completed_at')[:5]
    
    # Comparison statistikasi
    comparison_data = []
    for prev_result in previous_results:
        time_diff = None
        time_diff_abs = None
        if test_result.time_taken and prev_result.time_taken:
            time_diff = test_result.time_taken - prev_result.time_taken
            time_diff_abs = abs(time_diff)
        
        comparison_data.append({
            'result': prev_result,
            'score_diff': test_result.score - prev_result.score,
            'percentage_diff': round(test_result.percentage - prev_result.percentage, 1),
            'time_diff': time_diff,
            'time_diff_abs': time_diff_abs,
        })
    
    # Performance insights: question type bo'yicha aniqlik
    question_type_labels = dict(Question.QUESTION_TYPES)
    type_stats_map = {}
    all_questions = list(test_result.test.questions.all().order_by('order'))
    for q in all_questions:
        q_type = q.question_type or 'unknown'
        if q_type not in type_stats_map:
            type_stats_map[q_type] = {
                'question_type': q_type,
                'label': question_type_labels.get(q_type, q_type.replace('_', ' ').title()),
                'total': 0,
                'answered': 0,
                'correct': 0,
                'accuracy': 0.0,
            }
        type_stats_map[q_type]['total'] += 1
        ans = user_answers.get(q.id)
        if ans:
            type_stats_map[q_type]['answered'] += 1
            if ans.is_correct:
                type_stats_map[q_type]['correct'] += 1

    type_stats = []
    for item in type_stats_map.values():
        if item['answered'] > 0:
            item['accuracy'] = round((item['correct'] / item['answered']) * 100, 1)
        else:
            item['accuracy'] = 0.0
        # CSS width uchun lokalizatsiyasiz, xavfsiz qiymat
        item['accuracy_width'] = max(0, min(100, int(round(item['accuracy']))))
        type_stats.append(item)
    type_stats.sort(key=lambda x: x['accuracy'])

    weak_areas = [s for s in type_stats if s['answered'] > 0 and s['accuracy'] < 60][:3]
    top_strengths = sorted(type_stats, key=lambda x: x['accuracy'], reverse=True)[:3]

    seconds_per_question = 0
    if test_result.time_taken and test_result.total_questions:
        seconds_per_question = round(test_result.time_taken / max(test_result.total_questions, 1), 1)

    # Qisqa tavsiya generator
    tip_map = {
        'matching_headings': "Paragraphlarning asosiy g'oyasini 5-7 so'z bilan belgilash mashqini qiling.",
        'matching_features': "Ism/joy/obyektlarni alohida belgilab, keyin mapping qiling.",
        'matching_info': "Kalit so'zlarni sinonim bilan qidiring, to'g'ridan-to'g'ri so'zga bog'lanmang.",
        'true_false_not_given': "Gap ma'nosini 'True/False/Not Given'ga qat'iy ajratib tahlil qiling.",
        'yes_no_not_given': "Muallif fikri va faktni alohida baholang.",
        'fill_blank': "Word limitga qat'iy rioya qilib, imlo ustida ishlang.",
        'notes_completion': "Audio'da raqamlar, ism-sharif va joy nomlarini tez yozishga odatlaning.",
        'summary_completion': "Matnning mantiqiy oqimini (before/after) ushlashga e'tibor bering.",
        'sentence_completion': "Grammar mosligini (singular/plural, tense) tekshirib yozing.",
        'table_completion': "Jadval kategoriyalarini oldindan skanerlab oling.",
        'short_answer': "Qisqa va aniq javob yozing; keraksiz so'z qo'shmang.",
        'mcq': "Noto'g'ri variantlarni eliminatsiya usuli bilan qisqartiring.",
        'list_selection': "Kamida 2 marta qayta tekshirib, ortiqcha variantlarni olib tashlang.",
        'classification': "Har bir variantning xos belgisini bitta keyword bilan eslab boring.",
    }
    improvement_tips = [tip_map.get(w['question_type'], "Ushbu savol turida ko'proq amaliy test ishlang.") for w in weak_areas]

    test = test_result.test
    if getattr(test, 'variants_to_select', 1) == 2:
        ordered_questions = list(
            test.questions.order_by(Coalesce(F('variant'), Value(1)), 'order')
        )
    else:
        ordered_questions = list(test.questions.order_by('order'))
    question_display_num = {q.pk: i + 1 for i, q in enumerate(ordered_questions)}

    context = {
        'test_result': test_result,
        'user_answers': user_answers,
        'ordered_questions': ordered_questions,
        'question_display_num': question_display_num,
        'is_writing_test': test.test_type == 'writing',
        'can_retake': can_retake,
        'previous_results': previous_results,
        'comparison_data': comparison_data,
        'time_taken_hours': time_taken_hours,
        'time_taken_minutes': time_taken_minutes,
        'time_taken_seconds': time_taken_seconds,
        'type_stats': type_stats,
        'weak_areas': weak_areas,
        'top_strengths': top_strengths,
        'seconds_per_question': seconds_per_question,
        'improvement_tips': improvement_tips,
    }
    return render(request, 'core/tests/result.html', context)


@login_required
def profile(request):
    """Foydalanuvchi profili"""
    # Test natijalari (faqat tugallanganlar — jadval va hisobotlar uchun)
    test_results = UserTestResult.objects.filter(
        user=request.user,
        completed_at__isnull=False
    ).select_related('test', 'test__category').order_by('-completed_at')
    
    # Video progress
    video_progress = UserVideoProgress.objects.filter(
        user=request.user,
        watched=True
    ).select_related('video', 'video__category').order_by('-completed_at', '-last_watched_at')
    
    # Bookmarks
    bookmarks = Bookmark.objects.filter(user=request.user).select_related('video', 'test', 'test__category', 'video__category')[:10]
    
    # Study streak
    current_streak = StudyStreak.get_current_streak(request.user)
    recent_streaks = StudyStreak.objects.filter(user=request.user).order_by('-date')[:7]
    
    completed_results = test_results
    
    # Statistika (o'tgan testlar — har testning o'z passing_score bo'yicha)
    completed_list = list(completed_results.select_related('test'))
    stats = {
        'total_tests': len(completed_list),
        'total_videos': video_progress.count(),
        'average_score': completed_results.aggregate(Avg('percentage'))['percentage__avg'] or 0,
        'passed_tests': sum(1 for r in completed_list if r.is_passed()),
        'current_streak': current_streak,
        'total_bookmarks': Bookmark.objects.filter(user=request.user).count(),
    }
    
    # Hisobotlar kartochkalari uchun qisqa statistikalar
    now = timezone.now()
    week_start = now - timedelta(days=now.weekday())
    week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)
    week_end = week_start + timedelta(days=6, hours=23, minutes=59, seconds=59)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_day = monthrange(now.year, now.month)[1]
    month_end = now.replace(day=last_day, hour=23, minute=59, second=59, microsecond=0)
    
    weekly_tests = completed_results.filter(completed_at__gte=week_start, completed_at__lte=week_end)
    weekly_videos = video_progress.filter(completed_at__gte=week_start, completed_at__lte=week_end)
    monthly_tests = completed_results.filter(completed_at__gte=month_start, completed_at__lte=month_end)
    monthly_videos = video_progress.filter(completed_at__gte=month_start, completed_at__lte=month_end)
    
    analytics_summary = {
        'total_tests': stats['total_tests'],
        'avg_score': round(stats['average_score'], 1),
        'total_videos': stats['total_videos'],
    }
    weekly_summary_stats = {
        'total_tests': weekly_tests.count(),
        'avg_score': round(weekly_tests.aggregate(Avg('percentage'))['percentage__avg'] or 0, 1),
        'total_videos': weekly_videos.count(),
        'study_days': StudyStreak.objects.filter(user=request.user, date__gte=week_start.date(), date__lte=week_end.date()).count(),
    }
    monthly_summary_stats = {
        'total_tests': monthly_tests.count(),
        'avg_score': round(monthly_tests.aggregate(Avg('percentage'))['percentage__avg'] or 0, 1),
        'total_videos': monthly_videos.count(),
        'study_days': StudyStreak.objects.filter(user=request.user, date__gte=month_start.date(), date__lte=month_end.date()).count(),
    }
    
    context = {
        'test_results': test_results,
        'video_progress': video_progress,
        'bookmarks': bookmarks,
        'recent_streaks': recent_streaks,
        'stats': stats,
        'analytics_summary': analytics_summary,
        'weekly_summary_stats': weekly_summary_stats,
        'monthly_summary_stats': monthly_summary_stats,
    }
    return render(request, 'core/profile.html', context)


@login_required
def toggle_bookmark(request):
    """Bookmark qo'shish/olib tashlash"""
    if request.method == 'POST':
        video_id = request.POST.get('video_id')
        test_id = request.POST.get('test_id')
        
        if video_id:
            video = get_object_or_404(VideoLesson, pk=video_id)
            bookmark, created = Bookmark.objects.get_or_create(
                user=request.user,
                video=video
            )
            if not created:
                bookmark.delete()
                return JsonResponse({'bookmarked': False})
            return JsonResponse({'bookmarked': True})
        
        elif test_id:
            test = get_object_or_404(Test.objects.select_related('category'), pk=test_id)
            if not getattr(test.category, 'show_on_site', True):
                return JsonResponse({'error': 'Test mavjud emas.'}, status=404)
            bookmark, created = Bookmark.objects.get_or_create(
                user=request.user,
                test=test
            )
            if not created:
                bookmark.delete()
                return JsonResponse({'bookmarked': False})
            return JsonResponse({'bookmarked': True})
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
def leaderboard(request):
    """Eng yaxshi natijalar ro'yxati"""
    # Eng yaxshi natijalar (o'rtacha ball bo'yicha)
    top_users = UserTestResult.objects.values('user').annotate(
        avg_score=Avg('percentage'),
        total_tests=Count('id')
    ).filter(total_tests__gte=3).order_by('-avg_score')[:10]
    
    # User ID lar ro'yxati
    user_ids = [item['user'] for item in top_users]
    users = User.objects.filter(id__in=user_ids)
    user_dict = {user.id: user for user in users}
    
    # Leaderboard ma'lumotlari
    leaderboard_data = []
    for item in top_users:
        user = user_dict.get(item['user'])
        if user:
            leaderboard_data.append({
                'user': user,
                'avg_score': round(item['avg_score'], 1),
                'total_tests': item['total_tests']
            })
    
    # Foydalanuvchining o'z pozitsiyasi
    user_position = None
    user_avg = UserTestResult.objects.filter(user=request.user).aggregate(Avg('percentage'))['percentage__avg']
    if user_avg:
        all_users = UserTestResult.objects.values('user').annotate(
            avg_score=Avg('percentage'),
            total_tests=Count('id')
        ).filter(total_tests__gte=3).order_by('-avg_score')
        position = 1
        for item in all_users:
            if item['user'] == request.user.id:
                user_position = position
                break
            position += 1
    
    context = {
        'leaderboard': leaderboard_data,
        'user_position': user_position,
        'user_avg': round(user_avg, 1) if user_avg else 0,
    }
    return render(request, 'core/leaderboard.html', context)


@login_required
def statistics(request):
    """Batafsil statistika"""
    # Test natijalari statistikasi
    test_results = UserTestResult.objects.filter(user=request.user)
    
    # Kategoriya bo'yicha statistika
    category_stats = test_results.values('test__category__name').annotate(
        count=Count('id'),
        avg_score=Avg('percentage')
    ).order_by('-avg_score')
    
    # Test turi bo'yicha statistika
    type_stats = test_results.values('test__test_type').annotate(
        count=Count('id'),
        avg_score=Avg('percentage')
    )
    
    # Oylik statistika (oxirgi 6 oy)
    from datetime import timedelta
    from django.db.models.functions import TruncMonth
    monthly_stats = test_results.annotate(
        month=TruncMonth('completed_at')
    ).values('month').annotate(
        count=Count('id'),
        avg_score=Avg('percentage')
    ).order_by('-month')[:6]
    
    # Video statistika
    video_stats = UserVideoProgress.objects.filter(
        user=request.user,
        watched=True
    ).values('video__category__name').annotate(
        count=Count('id')
    )
    
    context = {
        'category_stats': category_stats,
        'type_stats': type_stats,
        'monthly_stats': monthly_stats,
        'video_stats': video_stats,
    }
    return render(request, 'core/statistics.html', context)


@login_required
def export_results(request):
    """Test natijalarini CSV'ga export qilish (faqat tugallangan natijalar)"""
    from django.http import HttpResponse
    import csv
    
    test_results = UserTestResult.objects.filter(
        user=request.user,
        completed_at__isnull=False
    ).select_related('test', 'test__category').order_by('-completed_at')
    
    filename = f"test_natijalari_{request.user.username}_{timezone.now().strftime('%Y%m%d')}.csv"
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # BOM qo'shish (Excel uchun)
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow(['Sana', 'Test', 'Kategoriya', 'Ball', 'Foiz', 'To\'g\'ri javoblar', 'Noto\'g\'ri javoblar', 'Holat'])
    
    for result in test_results:
        status = "O'tdi" if result.is_passed() else "O'tmadi"
        writer.writerow([
            result.completed_at.strftime('%d.%m.%Y %H:%M') if result.completed_at else '',
            result.test.title,
            result.test.category.name,
            f"{result.score}/{result.total_questions}",
            f"{result.percentage:.1f}%",
            result.correct_answers,
            result.wrong_answers,
            status
        ])
    
    return response


@login_required
def update_video_progress(request, pk):
    """Video progress yangilash"""
    if request.method == 'POST':
        video = get_object_or_404(VideoLesson, pk=pk, is_active=True)
        progress_percentage = int(request.POST.get('progress', 0))
        
        progress, created = UserVideoProgress.objects.get_or_create(
            user=request.user,
            video=video
        )
        
        progress.update_progress(progress_percentage)
        
        return JsonResponse({
            'success': True,
            'progress': progress.watch_percentage
        })
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
def add_video_note(request, pk):
    """Video eslatma qo'shish"""
    if request.method == 'POST':
        video = get_object_or_404(VideoLesson, pk=pk, is_active=True)
        note_text = request.POST.get('note_text', '').strip()
        timestamp = int(request.POST.get('timestamp', 0))
        
        if note_text:
            note = VideoNote.objects.create(
                user=request.user,
                video=video,
                note_text=note_text,
                timestamp=timestamp
            )
            
            return JsonResponse({
                'success': True,
                'note': {
                    'id': note.pk,
                    'text': note.note_text,
                    'timestamp': note.timestamp,
                    'timestamp_display': note.get_timestamp_display(),
                    'created_at': note.created_at.strftime('%d.%m.%Y %H:%M')
                }
            })
        
        return JsonResponse({'error': 'Eslatma matni bo\'sh bo\'lishi mumkin emas'}, status=400)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
def delete_video_note(request, note_id):
    """Video eslatma o'chirish"""
    if request.method == 'POST':
        note = get_object_or_404(VideoNote, pk=note_id, user=request.user)
        note.delete()
        
        return JsonResponse({'success': True})
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
def test_pause(request, pk):
    """Testni to'xtatish"""
    test = get_object_or_404(Test, pk=pk, is_active=True)
    test_result = get_object_or_404(
        UserTestResult,
        user=request.user,
        test=test,
        completed_at__isnull=True
    )
    
    test_result.pause_test()
    
    return JsonResponse({
        'success': True,
        'is_paused': True,
        'message': 'Test to\'xtatildi'
    })


@login_required
def test_resume(request, pk):
    """Testni davom ettirish"""
    test = get_object_or_404(Test, pk=pk, is_active=True)
    test_result = get_object_or_404(
        UserTestResult,
        user=request.user,
        test=test,
        completed_at__isnull=True
    )
    
    test_result.resume_test()
    
    return JsonResponse({
        'success': True,
        'is_paused': False,
        'message': 'Test davom etmoqda'
    })


@login_required
def test_update_time(request, pk):
    """Test vaqtini yangilash (AJAX)"""
    test = get_object_or_404(Test, pk=pk, is_active=True)
    test_result = get_object_or_404(
        UserTestResult,
        user=request.user,
        test=test,
        completed_at__isnull=True
    )
    
    if request.method == 'POST':
        elapsed_time = int(request.POST.get('elapsed_time', 0))
        timer_seconds_left = request.POST.get('timer_seconds_left')
        
        # Time tracking yangilash
        test_result.time_taken = elapsed_time
        if timer_seconds_left is not None:
            test_result.timer_seconds_left = int(timer_seconds_left)
        test_result.save(update_fields=['time_taken', 'timer_seconds_left'])
        
        return JsonResponse({
            'success': True,
            'elapsed_time': elapsed_time,
            'timer_seconds_left': test_result.timer_seconds_left,
            'is_paused': test_result.is_paused
        })
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
def rate_video(request, pk):
    """Video reyting qo'yish"""
    if request.method == 'POST':
        video = get_object_or_404(VideoLesson, pk=pk, is_active=True)
        
        try:
            rating_value = int(request.POST.get('rating', 0))
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Noto\'g\'ri reyting qiymati'}, status=400)
        
        if rating_value < 1 or rating_value > 5:
            return JsonResponse({'error': 'Reyting 1 dan 5 gacha bo\'lishi kerak'}, status=400)
        
        rating, created = VideoRating.objects.update_or_create(
            user=request.user,
            video=video,
            defaults={'rating': rating_value}
        )
        
        # Video obyektini refresh qilish - yangi reytinglarni olish uchun
        video.refresh_from_db()
        
        # O'rtacha reyting va jami reytinglarni hisoblash
        from django.db.models import Avg, Count
        rating_stats = VideoRating.objects.filter(video=video).aggregate(
            avg_rating=Avg('rating'),
            total=Count('id')
        )
        
        average_rating = round(rating_stats['avg_rating'], 1) if rating_stats['avg_rating'] else 0
        total_ratings = rating_stats['total'] or 0
        
        return JsonResponse({
            'success': True,
            'rating': rating.rating,
            'average_rating': average_rating,
            'total_ratings': total_ratings
        })
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
def add_video_comment(request, pk):
    """Video izoh qo'shish"""
    if request.method == 'POST':
        video = get_object_or_404(VideoLesson, pk=pk, is_active=True)
        comment_text = request.POST.get('comment_text', '').strip()
        parent_id = request.POST.get('parent_id')
        
        if not comment_text:
            return JsonResponse({'error': 'Izoh matni bo\'sh bo\'lishi mumkin emas'}, status=400)
        
        parent = None
        if parent_id:
            try:
                parent = VideoComment.objects.get(pk=parent_id, video=video)
            except VideoComment.DoesNotExist:
                return JsonResponse({'error': 'Noto\'g\'ri parent izoh'}, status=400)
        
        comment = VideoComment.objects.create(
            user=request.user,
            video=video,
            comment_text=comment_text,
            parent=parent
        )
        
        return JsonResponse({
            'success': True,
            'comment': {
                'id': comment.pk,
                'text': comment.comment_text,
                'user': comment.user.username,
                'user_first_name': comment.user.first_name or comment.user.username,
                'created_at': comment.created_at.strftime('%d.%m.%Y %H:%M'),
                'replies_count': 0,
                'parent_id': parent.pk if parent else None,
            }
        })
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
def delete_video_comment(request, comment_id):
    """Video izoh o'chirish"""
    if request.method == 'POST':
        comment = get_object_or_404(VideoComment, pk=comment_id, user=request.user)
        comment.delete()
        
        return JsonResponse({'success': True})
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
def create_playlist(request):
    """Playlist yaratish"""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        is_public = request.POST.get('is_public') == 'true'
        
        if not name:
            return JsonResponse({'error': 'Playlist nomi bo\'sh bo\'lishi mumkin emas'}, status=400)
        
        playlist = VideoPlaylist.objects.create(
            user=request.user,
            name=name,
            description=description,
            is_public=is_public
        )
        
        return JsonResponse({
            'success': True,
            'playlist': {
                'id': playlist.pk,
                'name': playlist.name,
                'description': playlist.description,
                'is_public': playlist.is_public,
            }
        })
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
def add_video_to_playlist(request, pk):
    """Videoni playlistga qo'shish"""
    if request.method == 'POST':
        video = get_object_or_404(VideoLesson, pk=pk, is_active=True)
        playlist_id = request.POST.get('playlist_id')
        
        if not playlist_id:
            return JsonResponse({'error': 'Playlist tanlanmagan'}, status=400)
        
        try:
            playlist = VideoPlaylist.objects.get(pk=playlist_id, user=request.user)
        except VideoPlaylist.DoesNotExist:
            return JsonResponse({'error': 'Playlist topilmadi'}, status=404)
        
        # Videoni playlistga qo'shish
        playlist_video, created = PlaylistVideo.objects.get_or_create(
            playlist=playlist,
            video=video,
            defaults={'order': playlist.videos_count + 1}
        )
        
        if not created:
            return JsonResponse({'error': 'Video allaqachon playlistda'}, status=400)
        
        return JsonResponse({
            'success': True,
            'message': 'Video playlistga qo\'shildi'
        })
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
def remove_video_from_playlist(request, pk):
    """Videoni playlistdan olib tashlash"""
    if request.method == 'POST':
        video = get_object_or_404(VideoLesson, pk=pk, is_active=True)
        playlist_id = request.POST.get('playlist_id')
        
        if not playlist_id:
            return JsonResponse({'error': 'Playlist tanlanmagan'}, status=400)
        
        try:
            playlist = VideoPlaylist.objects.get(pk=playlist_id, user=request.user)
        except VideoPlaylist.DoesNotExist:
            return JsonResponse({'error': 'Playlist topilmadi'}, status=404)
        
        PlaylistVideo.objects.filter(playlist=playlist, video=video).delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Video playlistdan olib tashlandi'
        })
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
def add_video_bookmark(request, pk):
    """Video ichida bookmark qo'yish (vaqt bilan)"""
    if request.method == 'POST':
        video = get_object_or_404(VideoLesson, pk=pk, is_active=True)
        timestamp = int(request.POST.get('timestamp', 0))
        
        # VideoNote ni bookmark sifatida ishlatish (timestamp bilan, lekin note_text bo'sh bo'lishi mumkin)
        # Yoki alohida VideoBookmark model yaratish mumkin, lekin VideoNote allaqachon timestamp bilan
        # Shuning uchun VideoNote ni bookmark sifatida ishlatamiz
        
        note, created = VideoNote.objects.get_or_create(
            user=request.user,
            video=video,
            timestamp=timestamp,
            defaults={'note_text': f'Bookmark at {timestamp}s'}
        )
        
        if not created:
            # Agar allaqachon mavjud bo'lsa, o'chirish (toggle)
            note.delete()
            return JsonResponse({
                'success': True,
                'bookmarked': False,
                'message': 'Bookmark olib tashlandi'
            })
        
        return JsonResponse({
            'success': True,
            'bookmarked': True,
            'bookmark': {
                'id': note.pk,
                'timestamp': note.timestamp,
                'timestamp_display': note.get_timestamp_display(),
            },
            'message': 'Bookmark qo\'shildi'
        })
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
def analytics(request):
    """Batafsil tahlillar"""
    # Vaqt oralig'i
    period = request.GET.get('period', 'all')  # all, week, month, year
    
    # Vaqt oralig'ini aniqlash
    now = timezone.now()
    if period == 'week':
        start_date = now - timedelta(days=7)
    elif period == 'month':
        start_date = now - timedelta(days=30)
    elif period == 'year':
        start_date = now - timedelta(days=365)
    else:
        start_date = None
    
    # Test natijalari
    test_results_query = UserTestResult.objects.filter(
        user=request.user,
        completed_at__isnull=False
    )
    if start_date:
        test_results_query = test_results_query.filter(completed_at__gte=start_date)
    
    test_results = test_results_query.select_related('test', 'test__category')
    
    # Asosiy statistika (o'tdi — har testning passing_score bo'yicha)
    test_results_list = list(test_results.select_related('test'))
    total_tests = len(test_results_list)
    avg_score = test_results.aggregate(Avg('percentage'))['percentage__avg'] or 0
    passed_tests = sum(1 for r in test_results_list if r.is_passed())
    failed_tests = total_tests - passed_tests
    
    # Kategoriya bo'yicha natijalar (o'tdi — har testning passing_score bo'yicha)
    category_performance = test_results.values('test__category__name').annotate(
        count=Count('id'),
        avg_score=Avg('percentage'),
        passed=Count('id', filter=Q(percentage__gte=F('test__passing_score'))),
        failed=Count('id', filter=Q(percentage__lt=F('test__passing_score')))
    ).order_by('-avg_score')
    
    # Test turlari bo'yicha natijalar
    test_type_performance = test_results.values('test__test_type').annotate(
        count=Count('id'),
        avg_score=Avg('percentage')
    ).order_by('-avg_score')
    
    # Qiyinlik darajasi bo'yicha natijalar
    difficulty_performance = test_results.values('test__difficulty').annotate(
        count=Count('id'),
        avg_score=Avg('percentage')
    ).order_by('-avg_score')
    
    # Vaqt bo'yicha progress (kunlik)
    from django.db.models.functions import TruncDate
    daily_progress = test_results.annotate(
        day=TruncDate('completed_at')
    ).values('day').annotate(
        count=Count('id'),
        avg_score=Avg('percentage')
    ).order_by('day')[:30]
    
    # Eng yaxshi va eng yomon natijalar
    best_result = test_results.order_by('-percentage').first()
    worst_result = test_results.order_by('percentage').first()
    
    # O'rtacha vaqt (soniya) va o'qilishi oson format
    avg_time_sec = test_results.filter(time_taken__gt=0).aggregate(Avg('time_taken'))['time_taken__avg'] or 0
    avg_time = int(avg_time_sec)
    if avg_time >= 3600:
        avg_time_display = f"{avg_time // 3600} soat {(avg_time % 3600) // 60} d"
    elif avg_time >= 60:
        avg_time_display = f"{avg_time // 60} d"
    else:
        avg_time_display = f"{avg_time} s"
    
    # Video statistikasi (tanlangan davr uchun)
    video_progress_query = UserVideoProgress.objects.filter(user=request.user, watched=True)
    if start_date:
        video_progress_query = video_progress_query.filter(completed_at__gte=start_date)
    total_videos_watched = video_progress_query.count()
    
    # Grafiklar uchun ma'lumotlar
    daily_labels = []
    for item in daily_progress:
        if item['day']:
            # Agar day date object bo'lsa
            if hasattr(item['day'], 'strftime'):
                daily_labels.append(item['day'].strftime('%d.%m'))
            # Agar day string bo'lsa
            elif isinstance(item['day'], str):
                try:
                    from datetime import datetime
                    date_obj = datetime.strptime(item['day'], '%Y-%m-%d').date()
                    daily_labels.append(date_obj.strftime('%d.%m'))
                except:
                    daily_labels.append(str(item['day'])[:5])
            else:
                daily_labels.append(str(item['day'])[:5])
        else:
            daily_labels.append('')
    
    chart_data = {
        'category_labels': [item['test__category__name'] for item in category_performance],
        'category_scores': [round(item['avg_score'], 1) for item in category_performance],
        'category_counts': [item['count'] for item in category_performance],
        'daily_labels': daily_labels,
        'daily_scores': [round(item['avg_score'], 1) for item in daily_progress],
        'daily_counts': [item['count'] for item in daily_progress],
    }
    
    context = {
        'period': period,
        'total_tests': total_tests,
        'avg_score': round(avg_score, 1),
        'passed_tests': passed_tests,
        'failed_tests': failed_tests,
        'total_videos_watched': total_videos_watched,
        'category_performance': category_performance,
        'test_type_performance': test_type_performance,
        'difficulty_performance': difficulty_performance,
        'daily_progress': daily_progress,
        'best_result': best_result,
        'worst_result': worst_result,
        'avg_time': avg_time,
        'avg_time_display': avg_time_display,
        'chart_data': json.dumps(chart_data),
    }
    return render(request, 'core/analytics.html', context)


@login_required
def export_to_excel(request):
    """Test natijalarini Excel'ga export qilish"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'Excel export funksiyasi ishlamayapti. Iltimos, openpyxl paketini o\'rnating.')
        return redirect('core:profile')
    
    # Test natijalari
    test_results = UserTestResult.objects.filter(
        user=request.user,
        completed_at__isnull=False
    ).select_related('test', 'test__category').order_by('-completed_at')
    
    # Workbook yaratish
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Natijalari"
    
    # Header
    headers = ['Sana', 'Test', 'Kategoriya', 'Ball', 'Foiz', "To'g'ri", "Noto'g'ri", 'Vaqt (soniya)', 'Holat']
    ws.append(headers)
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data
    for result in test_results:
        status = "O'tdi" if result.is_passed() else "O'tmadi"
        time_str = f"{result.time_taken // 60}:{result.time_taken % 60:02d}" if result.time_taken else "-"
        ws.append([
            result.completed_at.strftime('%d.%m.%Y %H:%M'),
            result.test.title,
            result.test.category.name,
            f"{result.score}/{result.total_questions}",
            f"{result.percentage:.1f}%",
            result.correct_answers,
            result.wrong_answers,
            time_str,
            status
        ])
    
    # Column width
    column_widths = [18, 40, 20, 12, 10, 10, 10, 15, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="test_natijalari_{request.user.username}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
def weekly_summary(request):
    """Haftalik xulosa"""
    now = timezone.now()
    week_start = now - timedelta(days=now.weekday())
    week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)
    week_end = week_start + timedelta(days=6, hours=23, minutes=59, seconds=59)
    
    # Test natijalari
    test_results = UserTestResult.objects.filter(
        user=request.user,
        completed_at__gte=week_start,
        completed_at__lte=week_end,
        completed_at__isnull=False
    ).select_related('test', 'test__category')
    
    # Video progress
    video_progress = UserVideoProgress.objects.filter(
        user=request.user,
        completed_at__gte=week_start,
        completed_at__lte=week_end,
        watched=True
    ).select_related('video', 'video__category')
    
    # Statistika
    total_tests = test_results.count()
    avg_score = test_results.aggregate(Avg('percentage'))['percentage__avg'] or 0
    total_videos = video_progress.count()
    study_days = StudyStreak.objects.filter(
        user=request.user,
        date__gte=week_start.date(),
        date__lte=week_end.date()
    ).count()
    
    context = {
        'week_start': week_start,
        'week_end': week_end,
        'test_results': test_results,
        'video_progress': video_progress,
        'total_tests': total_tests,
        'avg_score': round(avg_score, 1),
        'total_videos': total_videos,
        'study_days': study_days,
    }
    return render(request, 'core/weekly_summary.html', context)


@login_required
def monthly_report(request):
    """Oylik hisobot"""
    now = timezone.now()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_day = monthrange(now.year, now.month)[1]
    month_end = now.replace(day=last_day, hour=23, minute=59, second=59, microsecond=0)
    
    # Test natijalari
    test_results = UserTestResult.objects.filter(
        user=request.user,
        completed_at__gte=month_start,
        completed_at__lte=month_end,
        completed_at__isnull=False
    ).select_related('test', 'test__category')
    
    # Video progress
    video_progress = UserVideoProgress.objects.filter(
        user=request.user,
        completed_at__gte=month_start,
        completed_at__lte=month_end,
        watched=True
    ).select_related('video', 'video__category')
    
    # Kategoriya bo'yicha statistika
    category_stats = test_results.values('test__category__name').annotate(
        count=Count('id'),
        avg_score=Avg('percentage'),
        passed=Count('id', filter=Q(percentage__gte=60))
    ).order_by('-count')
    
    # Statistika
    total_tests = test_results.count()
    avg_score = test_results.aggregate(Avg('percentage'))['percentage__avg'] or 0
    total_videos = video_progress.count()
    study_days = StudyStreak.objects.filter(
        user=request.user,
        date__gte=month_start.date(),
        date__lte=month_end.date()
    ).count()
    
    context = {
        'month_start': month_start,
        'month_end': month_end,
        'test_results': test_results,
        'video_progress': video_progress,
        'category_stats': category_stats,
        'total_tests': total_tests,
        'avg_score': round(avg_score, 1),
        'total_videos': total_videos,
        'study_days': study_days,
    }
    return render(request, 'core/monthly_report.html', context)


@staff_member_required
def admin_toliq_yoriqnoma(request):
    """Admin uchun bitta sahifada to'liq yo'riqnoma — Test, Part, Savol qo'shish."""
    return render(request, 'admin/core/toliq_yoriqnoma.html')
